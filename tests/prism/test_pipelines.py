"""Tests for pipeline config generation."""
import os
import copy
from tempfile import NamedTemporaryFile
from typing import List, Tuple
import openpyxl
import yaml

import pytest
from cidc_schemas.prism.constants import SUPPORTED_SHIPPING_MANIFESTS

from cidc_schemas.template import Template
from cidc_schemas.template_reader import XlTemplateReader
from cidc_schemas.prism import core, pipelines, merger

from ..constants import TEMPLATE_EXAMPLES_DIR
from ..test_templates import (
    template_set,
    template,
    template_example,
    template_example_xlsx_path,
)
from .cidc_test_data import get_test_trial


@pytest.fixture(scope="session")
def prismify_result(template, template_example):

    # tear down
    core._encrypt_hmac = None
    # and set up
    core.set_prism_encrypt_key("test")

    prism_patch, file_maps, errs = core.prismify(template_example, template)
    assert not errs, "\n".join([str(e) for e in errs])
    return prism_patch, file_maps, errs


def prism_patch_stage_artifacts(
    prismify_result, template_type
) -> Tuple[dict, List[Tuple[dict, dict]]]:

    prism_patch, prism_fmap, _ = prismify_result
    patch_copy_4_artifacts = copy.deepcopy(prism_patch)

    patch_copy_4_artifacts = merger.merge_artifacts(
        patch_copy_4_artifacts,
        [
            merger.ArtifactInfo(
                artifact_uuid=fmap_entry.upload_placeholder,
                object_url=fmap_entry.gs_key,
                upload_type=template_type,
                file_size_bytes=i,
                uploaded_timestamp="01/01/2001",
                md5_hash=f"hash_{i}",
            )
            for i, fmap_entry in enumerate(prism_fmap)
        ],
    )[0]

    return patch_copy_4_artifacts


def stage_assay_for_analysis(template_type) -> Tuple[dict, List[Tuple[dict, dict]]]:
    """
    Simulates an initial assay upload by prismifying the initial assay template object.
    """

    staging_map = {
        "cytof_analysis": "cytof",
        "tumor_normal_pairing": "wes_fastq",
    }

    if not template_type in staging_map:
        return {}

    prelim_assay = staging_map[template_type]

    return stage_assay(template_type=prelim_assay)


def stage_assay(template_type: str) -> Tuple[dict, List[Tuple[dict, dict]]]:
    preassay_xlsx_path = os.path.join(
        TEMPLATE_EXAMPLES_DIR, template_type + "_template.xlsx"
    )
    preassay_xlsx, _ = XlTemplateReader.from_excel(preassay_xlsx_path)
    preassay_template = Template.from_type(template_type)
    prism_res = core.prismify(preassay_xlsx, preassay_template)

    return prism_patch_stage_artifacts(prism_res, template_type)


def test_WES_pipeline_config_generation_after_prismify(prismify_result, template):

    if not (template.type.startswith("wes_") or "pair" in template.type):
        return

    # Test that the config generator blocks disallowed upload types
    upload_type = "foo"
    with pytest.raises(NotImplementedError, match=f"Not supported type:{upload_type}"):
        pipelines._Wes_pipeline_config(upload_type)

    full_ct = get_test_trial(
        [
            "CTTTPP111.00",
            "CTTTPP121.00",
            "CTTTPP122.00",
            "CTTTPP123.00",
            "CTTTPP124.00",
            "CTTTPP211.00",
            "CTTTPP212.00",
            "CTTTPP213.00",
            "CTTTPP214.00",
            "CTTTPP311.00",
            "CTTTPP312.00",
            "CTTTPP313.00",
            "CTTTPP411.00",
            "CTTTPP412.00",
            "CTTTPP413.00",
            "CTTTPP511.00",
        ],
        allowed_collection_event_names=[
            "Not_reported",
            "Baseline",
            "On_Treatment",
            "Week_1",
        ],
        assays={"wes": []},
    )
    # manually modify json's to add tumor / normal definitions for WES
    # these are normally loaded from the shipping manifests
    if "wes" in template.type or template.type == "tumor_normal_pairing":
        for partic in full_ct["participants"]:
            partic_id = partic["cimac_participant_id"]
            if partic_id == "CTTTPP1":
                # these are paired in tumor_normal_pairing
                partic["samples"][0]["processed_sample_derivative"] = "Tumor DNA"
                partic["samples"][1]["processed_sample_derivative"] = "Germline DNA"

                # test default to tumor if not specified
                partic["samples"][2]["collection_event_name"] = "Baseline"

                # test deduplication of normals by collection_event_name
                for n in (3, 4):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                    partic["samples"][n]["collection_event_name"] = "Baseline"
            elif partic_id == "CTTTPP2":
                # test 2 tumor samples on treatment and only 1 normal on treatment
                for n in (0, 1):
                    partic["samples"][n]["processed_sample_derivative"] = "Tumor DNA"
                    partic["samples"][n]["collection_event_name"] = "On_Treatment"
                for n in (2, 3):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                partic["samples"][2]["collection_event_name"] = "On_Treatment"
                partic["samples"][3]["collection_event_name"] = "Baseline"
            elif partic_id == "CTTTPP3":
                # test 1 tumor sample not reported and one of normal samples baseline
                partic["samples"][0]["processed_sample_derivative"] = "Tumor DNA"
                for n in (1, 2):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                partic["samples"][1]["collection_event_name"] = "On_Treatment"
                partic["samples"][2]["collection_event_name"] = "Baseline"
            elif partic_id == "CTTTPP4":
                # test 1 tumor sample not reported and 2 normals other collection events
                partic["samples"][0]["processed_sample_derivative"] = "Tumor DNA"
                for n in (1, 2):
                    partic["samples"][n]["processed_sample_derivative"] = "Germline DNA"
                partic["samples"][1]["collection_event_name"] = "On_Treatment"
                partic["samples"][2]["collection_event_name"] = "Week_1"
            elif partic_id == "CTTTPP5":
                # test 1 tumor sample with no paired normal
                partic["samples"][0]["processed_sample_derivative"] = "Tumor DNA"
                partic["samples"][0]["collection_event_name"] = "On_Treatment"

    patch_with_artifacts = prism_patch_stage_artifacts(prismify_result, template.type)

    # if it's an analysis - we need to merge corresponding preliminary assay first
    prelim_assay = stage_assay_for_analysis(template.type)
    if prelim_assay:
        full_ct, errs = merger.merge_clinical_trial_metadata(prelim_assay, full_ct)
        assert 0 == len(errs), str(errs)

    full_ct, errs = merger.merge_clinical_trial_metadata(patch_with_artifacts, full_ct)
    assert 0 == len(errs), str(errs)

    res = pipelines.generate_analysis_configs_from_upload_patch(
        full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
    )

    pairing_filename = full_ct["protocol_identifier"] + "_pairing.csv"
    # wes_bam
    if template.type == "wes_bam":
        # 2 config, tumor-only for both samples
        # 2 template, wes_tumor_only_analysis for both
        # 1 pairing file
        assert len(res) == 5
        assert (
            res[pairing_filename]
            == "protocol_identifier,test_prism_trial_id\ntumor,tumor_collection_event,normal,normal_collection_event\nCTTTPP111.00,Not_reported,CTTTPP121.00,Not_reported"
        )
    elif template.type == "tumor_normal_pairing":
        # 1 config, tumor/normal, config for CTTTPP122.00 was generated on wes_fastq upload
        # 1 template, wes_analysis for pair
        # 1 pairing file
        assert len(res) == 3
        assert (
            res[pairing_filename] == "protocol_identifier,test_prism_trial_id\n"
            "tumor,tumor_collection_event,normal,normal_collection_event\n"
            "CTTTPP111.00,Not_reported,CTTTPP121.00,Not_reported\n"
            "CTTTPP122.00,Baseline,CTTTPP123.00,Baseline\n"
            "CTTTPP211.00,On_Treatment,CTTTPP213.00,On_Treatment\n"
            "CTTTPP212.00,On_Treatment,CTTTPP213.00,On_Treatment\n"
            ",,CTTTPP214.00,Baseline\n"
            "CTTTPP311.00,Not_reported,CTTTPP313.00,Baseline\n"
            ",,CTTTPP312.00,On_Treatment\n"
            "CTTTPP411.00,Not_reported,,\n"
            ",,CTTTPP412.00,On_Treatment\n"
            ",,CTTTPP413.00,Week_1\n"
            "CTTTPP511.00,On_Treatment,,"
        )
    elif template.type == "wes_fastq":
        # 16 configs, tumor-only for all samples
        # 16 wes_tumor_only_analysis for all
        # 1 pairing file
        assert len(res) == 33
        assert (
            res[pairing_filename] == "protocol_identifier,test_prism_trial_id\n"
            "tumor,tumor_collection_event,normal,normal_collection_event\n"
            "CTTTPP111.00,Not_reported,CTTTPP121.00,Not_reported\n"
            "CTTTPP122.00,Baseline,CTTTPP123.00,Baseline\n"
            "CTTTPP211.00,On_Treatment,CTTTPP213.00,On_Treatment\n"
            "CTTTPP212.00,On_Treatment,CTTTPP213.00,On_Treatment\n"
            ",,CTTTPP214.00,Baseline\n"
            "CTTTPP311.00,Not_reported,CTTTPP313.00,Baseline\n"
            ",,CTTTPP312.00,On_Treatment\n"
            "CTTTPP411.00,Not_reported,,\n"
            ",,CTTTPP412.00,On_Treatment\n"
            ",,CTTTPP413.00,Week_1\n"
            "CTTTPP511.00,On_Treatment,,"
        )
    else:  # where we don't expect to have configs
        assert res == {}
        return

    for fname, conf in res.items():
        if fname.endswith("yaml"):
            conf = yaml.load(conf, Loader=yaml.FullLoader)
            assert len(conf["metasheet"]) == 1  # one run

            if "pair" in template.type:
                assert len(conf["samples"]) in [1, 2]
                assert conf["instance_name"] == (
                    "ctttpp111-00"  # run_id from sheet
                    if len(conf["samples"]) == 2  # tumor/normal
                    else "ctttpp122-00"  # tumor id for tumor-only
                )  # run ID but lowercase & hyphenated

            else:
                assert len(conf["samples"]) == 1  # tumor only
                assert conf["instance_name"] in [
                    "ctttpp111-00",
                    "ctttpp121-00",
                    "ctttpp122-00",
                    "ctttpp123-00",
                    "ctttpp124-00",
                    "ctttpp211-00",
                    "ctttpp212-00",
                    "ctttpp213-00",
                    "ctttpp214-00",
                    "ctttpp311-00",
                    "ctttpp312-00",
                    "ctttpp313-00",
                    "ctttpp411-00",
                    "ctttpp412-00",
                    "ctttpp413-00",
                    "ctttpp511-00",
                ]  # CIMAC ID but lowercase & hypenated

            for sample in conf["samples"].values():
                assert len(sample) > 0  # at least one data file per sample
                assert all("my-biofx-bucket" in f for f in sample)
                assert all(f.endswith(".fastq.gz") for f in sample) or all(
                    f.endswith(".bam") for f in sample
                )

        elif fname.endswith("xlsx"):
            # openpyxl needs to file to have an .xlsx extension to open it
            with NamedTemporaryFile(suffix=".xlsx") as tmp:
                tmp.write(conf)
                tmp.seek(0)
                wb = openpyxl.load_workbook(tmp.name, data_only=True)

            if "WES Analysis" in wb.sheetnames:
                sht = wb["WES Analysis"]
            elif "WES tumor-only Analysis" in wb.sheetnames:
                sht = wb["WES tumor-only Analysis"]
            else:
                assert (
                    False
                ), f"Attached xlsx doesn't have right worksheets: {wb.sheetnames}"

            assert sht["C2"].value == "test_prism_trial_id"
            assert sht["C3"].value == pipelines.BIOFX_WES_ANALYSIS_FOLDER
            assert sht["B7"].value  # run name
            assert sht["C7"].value  # first id

            if sht.title == "WES Analysis":
                assert sht["D7"].value  # second id


def test_RNAseq_pipeline_config_generation_after_prismify(prismify_result, template):

    if not template.type.startswith("rna_"):
        return

    full_ct = get_test_trial(
        ["CTTTPP111.00", "CTTTPP121.00", "CTTTPP122.00", "CTTTPP123.00"],
        assays={"rna": []},
    )
    patch_with_artifacts = prism_patch_stage_artifacts(prismify_result, template.type)

    # if it's an analysis - we need to merge corresponding preliminary assay first
    prelim_assay = stage_assay_for_analysis(template.type)
    if prelim_assay:
        full_ct, errs = merger.merge_clinical_trial_metadata(prelim_assay, full_ct)
        assert 0 == len(errs)

    full_ct, errs = merger.merge_clinical_trial_metadata(patch_with_artifacts, full_ct)

    assert 0 == len(errs), "\n".join(errs)

    res = pipelines.generate_analysis_configs_from_upload_patch(
        full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
    )

    # where we don't expect to have configs
    if not template.type in pipelines._ANALYSIS_CONF_GENERATORS:
        assert res == {}
        return

    if template.type in ["rna_fastq", "rna_bam"]:
        # one config with all samples from one participant in one example .xlsx
        # plus one metasheet.csv
        assert len(res) == 1 + 1

    else:
        assert False, f"Unexpected RNAseq template test {template.type}"

    for fname, fcontent in res.items():

        if not fname.endswith(".yaml"):
            assert fname.endswith(".csv")

            assert (
                fcontent
                == "cimac_id,cimac_participant_id,collection_event_name,type_of_sample,processed_sample_derivative"
                "\r\nCTTTPP122.00,CTTTPP1,Not_reported,Not Reported,"
                "\r\nCTTTPP123.00,CTTTPP1,Not_reported,Not Reported,"
            )

        else:

            conf = yaml.load(fcontent, Loader=yaml.FullLoader)

            assert len(conf["runs"]) == 2  # two runs for two samples in example .xlsx

            assert len(conf["samples"]) == 2  # two samples in example .xlsx
            for sample in conf["samples"].values():
                assert len(sample) > 0  # at lease one data file per sample
                assert all("my-biofx-bucket" in f for f in sample)
                assert all(f.endswith(".fastq.gz") for f in sample) or all(
                    f.endswith(".bam") for f in sample
                )


def test_shipping_manifest_new_participants_after_prismify(prismify_result, template):

    if not template.type in SUPPORTED_SHIPPING_MANIFESTS:
        return

    base_ct = get_test_trial(
        allowed_cohort_names=["Arm_A", "Arm_Z"],
        allowed_collection_event_names=[
            "Baseline",
            "Pre_Day_1_Cycle_2",
        ],
    )

    patch_with_artifacts = prism_patch_stage_artifacts(prismify_result, template.type)
    full_ct, errs = merger.merge_clinical_trial_metadata(patch_with_artifacts, base_ct)
    assert 0 == len(errs), "\n".join(errs)

    # test returns all participants on first upload
    res = pipelines.generate_analysis_configs_from_upload_patch(
        full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
    )
    assert len(res) == 1

    expected_answer = {
        "h_and_e": ["CTTTP08"],
        "microbiome_dna": ["CTTTP08", "CTTTP09"],
        "normal_blood_dna": ["CTTTP01", "CTTTP02"],
        "normal_tissue_dna": ["CTTTP03", "CTTTP04"],
        "pbmc": ["CTTTP01", "CTTTP02"],
        "plasma": ["CTTTP01", "CTTTP02"],
        "tissue_slide": ["CTTTP08", "CTTTP09"],
        "tumor_tissue_dna": ["CTTTP05", "CTTTP06"],
        "tumor_tissue_rna": ["CTTTP05", "CTTTP06"],
    }
    assert res["new_participants.txt"].split("\n") == expected_answer[template.type]

    # test ONLY new participants on subset
    if template.type == "h_and_e":
        patch_with_artifacts = stage_assay(template_type="microbiome_dna")
        full_ct, errs = merger.merge_clinical_trial_metadata(
            patch_with_artifacts, full_ct
        )
        assert 0 == len(errs), "\n".join(errs)

        res = pipelines.generate_analysis_configs_from_upload_patch(
            full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
        )
        assert len(res) == 1

        # doesn't return the one from h_and_e
        assert (
            res["new_participants.txt"].split("\n")
            == expected_answer["microbiome_dna"][1:]
        )

    # test doesn't return if no new particpants
    if template.type == "pbmc":
        patch_with_artifacts = stage_assay(template_type="plasma")
        full_ct, errs = merger.merge_clinical_trial_metadata(
            patch_with_artifacts, full_ct
        )
        assert 0 == len(errs), "\n".join(errs)
        full_ct, errs = merger.merge_clinical_trial_metadata(
            patch_with_artifacts, full_ct
        )
        assert 0 == len(errs), "\n".join(errs)

        res = pipelines.generate_analysis_configs_from_upload_patch(
            full_ct, patch_with_artifacts, template.type, "my-biofx-bucket"
        )
        assert res == dict()
